import openpyxl

wb = openpyxl.load_workbook("plantilla.xlsx")

print(wb.sheetnames)
print(wb["Hoja1"])

# devolver hoja activa

hoja = wb.active

print(hoja)

# crear una hoja

wb.create_sheet("Hoja2")
print(wb.get_sheet_names())


hoja2 = wb["Hoja2"]
hoja2.title = "Hoja2_Nuevo"


print(
    hoja.max_row,
    hoja.max_column

)

celda = hoja["A1"]
celda.value = "Nombre Completo"
print(celda.value)

celda2 = hoja.cell(row=2, column=1)
print(
    celda2.value,
    celda2.row,
    celda2.column,
    celda2.coordinate
)

for fila in range(1, hoja.max_row + 1):
    for columna in range(1, hoja.max_column + 1):
        celda = hoja.cell(row=fila, column=columna)
        print(celda.value,
              celda.coordinate)


columna = hoja["A"]
fila = hoja["1"]
print(columna, fila)

hoja.append([1, 2, 3])
print(hoja.rows)

hoja.delete_rows(1, 1)


wb.save("plantilla_nueva.xlsx")
